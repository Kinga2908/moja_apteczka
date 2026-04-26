from django.shortcuts import render, redirect
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from .forms import LekForm, PrzyjęcieForm, UserProfileForm, RejestracjaForm
from .models import Lek, PrzyjecieLeku, UserProfile
import csv
import io
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def rejestracja(request):
    if request.method == 'POST':
        form = RejestracjaForm(request.POST)
        if form.is_valid():
            form.save()                    # zapisz usera
            # login(request, user)        # ← usuń tę linię
            return redirect('logowanie')  # ← przekieruj do logowania
    else:
        form = RejestracjaForm()
    return render(request, 'leki/formularz.html', {
        'form': form,
        'tytul': 'Rejestracja'
    })


def logowanie(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('strona_glowna')
    else:
        form = AuthenticationForm()
    return render(request, 'leki/formularz.html', {
        'form': form,
        'tytul': 'Logowanie'
    })


def wylogowanie(request):
    logout(request)
    return redirect('logowanie')


@login_required
def strona_glowna(request):
    from datetime import date, timedelta
    offset = int(request.GET.get('tydzien', 0)) 
    
    leki = Lek.objects.filter(uzytkownik=request.user)
    przyjecia = PrzyjecieLeku.objects.filter(uzytkownik=request.user)
    
    return render(request, 'leki/strona_glowna.html', {
        'leki': leki,
        'przyjecia': przyjecia,
        'tydzien_offset': offset,
    })


@login_required
def dodaj_lek(request):
    if request.method == 'POST':
        form = LekForm(request.POST)
        if form.is_valid():
            lek = form.save(commit=False)
            lek.uzytkownik = request.user 
            lek.save()
            form.save_m2m()              
            return redirect('strona_glowna')
    else:
        form = LekForm()
    return render(request, 'leki/formularz.html', {
        'form': form,
        'tytul': 'Dodaj nowy lek'
    })


@login_required
def dodaj_przyjecie(request):
    if request.method == 'POST':
        form = PrzyjęcieForm(request.POST)
        if form.is_valid():
            przyjecie = form.save(commit=False)
            przyjecie.uzytkownik = request.user
            przyjecie.save()
            form.save_m2m()
            return redirect('strona_glowna')
    else:
        form = PrzyjęcieForm()
    return render(request, 'leki/formularz.html', {
        'form': form,
        'tytul': 'Zarejestruj przyjęcie leku'
    })


@login_required
def edytuj_profil(request):
    profil, created = UserProfile.objects.get_or_create(uzytkownik=request.user)
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=profil)
        if form.is_valid():
            form.save()
            return redirect('strona_glowna')
    else:
        form = UserProfileForm(instance=profil)
    return render(request, 'leki/formularz.html', {
        'form': form,
        'tytul': 'Edytuj profil'
    })

@login_required
def edytuj_przyjecie(request, pk):
    przyjecie = PrzyjecieLeku.objects.get(pk=pk, uzytkownik=request.user)
    if request.method == 'POST':
        form = PrzyjęcieForm(request.POST, instance=przyjecie)
        if form.is_valid():
            form.save()
            return redirect('strona_glowna')
    else:
        form = PrzyjęcieForm(instance=przyjecie)
    return render(request, 'leki/formularz.html', {
        'form': form,
        'tytul': 'Edytuj przyjęcie'
    })

@login_required
def zmien_status(request, pk):
    przyjecie = PrzyjecieLeku.objects.get(pk=pk, uzytkownik=request.user)
    if przyjecie.status == 'zazyte':
        przyjecie.status = 'niezazyte'
    else:
        przyjecie.status = 'zazyte'
    przyjecie.save()
    return redirect('strona_glowna')


@login_required
def eksport_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="apteczka.csv"'
    response.write('\ufeff')  # BOM dla Excela

    writer = csv.writer(response)
    writer.writerow(['Lek', 'Substancja aktywna', 'Data przyjęcia', 'Status', 'Samopoczucie'])

    przyjecia = PrzyjecieLeku.objects.filter(uzytkownik=request.user).select_related('lek')
    for p in przyjecia:
        writer.writerow([
            p.lek.nazwa,
            p.lek.substancja_aktywna,
            p.data_godzina.strftime('%Y-%m-%d %H:%M'),
            p.get_status_display(),
            p.notatka_samopoczucia,
        ])

    return response


@login_required
def eksport_xlsx(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Apteczka"

    # Nagłówki
    naglowki = ['Lek', 'Substancja aktywna', 'Data przyjęcia', 'Status', 'Samopoczucie']
    ws.append(naglowki)

    # Styl nagłówków
    green_fill = PatternFill(start_color='2C7A4B', end_color='2C7A4B', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = green_fill
        cell.alignment = Alignment(horizontal='center')

    # Dane
    przyjecia = PrzyjecieLeku.objects.filter(uzytkownik=request.user).select_related('lek')
    for p in przyjecia:
        ws.append([
            p.lek.nazwa,
            p.lek.substancja_aktywna,
            p.data_godzina.strftime('%Y-%m-%d %H:%M'),
            p.get_status_display(),
            p.notatka_samopoczucia,
        ])

    # Szerokości kolumn
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="apteczka.xlsx"'
    wb.save(response)
    return response


@login_required
def wykres_png(request):
    """Serwuje dynamiczny wykres przyjęć z ostatnich 14 dni jako plik PNG."""
    przyjecia = PrzyjecieLeku.objects.filter(uzytkownik=request.user)

    today = timezone.now().date()
    dni = [(today - timedelta(days=i)) for i in range(13, -1, -1)]
    counts = {d: 0 for d in dni}
    for p in przyjecia:
        d = p.data_godzina.date()
        if d in counts:
            counts[d] += 1

    labels = [d.strftime('%d.%m') for d in dni]
    values = [counts[d] for d in dni]

    fig, ax = plt.subplots(figsize=(10, 4))
    bars = ax.bar(labels, values, color='#2C7A4B', edgecolor='#1a4d2e', linewidth=0.8)

    # Wartości nad słupkami
    for bar, val in zip(bars, values):
        if val > 0:
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.05,
                    str(val), ha='center', va='bottom', fontsize=9, color='#1a4d2e')

    ax.set_title('Przyjęcia leków — ostatnie 14 dni', fontsize=13, fontweight='bold', color='#1a4d2e')
    ax.set_ylabel('Liczba przyjęć')
    ax.set_ylim(0, max(max(values) + 1, 5))
    ax.set_facecolor('#f8fdf9')
    fig.patch.set_facecolor('#f8fdf9')
    ax.tick_params(axis='x', rotation=45, labelsize=8)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=120)
    plt.close(fig)
    buf.seek(0)

    return HttpResponse(buf, content_type='image/png')


@login_required
def import_csv(request):
    from .forms import ImportCSVForm
    bledy = []
    sukces = 0

    if request.method == 'POST':
        form = ImportCSVForm(request.POST, request.FILES)
        if form.is_valid():
            plik = request.FILES['plik']

            # Sprawdź rozszerzenie
            if not plik.name.endswith('.csv'):
                bledy.append('Plik musi mieć rozszerzenie .csv')
            else:
                try:
                    decoded = plik.read().decode('utf-8-sig')  # obsługuje BOM
                    reader = csv.DictReader(io.StringIO(decoded))

                    wymagane = {'nazwa', 'substancja_aktywna', 'instrukcja'}
                    if not wymagane.issubset(set(reader.fieldnames or [])):
                        bledy.append(f'Brakuje kolumn. Wymagane: {", ".join(wymagane)}')
                    else:
                        for i, row in enumerate(reader, start=2):
                            nazwa = row.get('nazwa', '').strip()
                            substancja = row.get('substancja_aktywna', '').strip()
                            instrukcja = row.get('instrukcja', '').strip()

                            if not nazwa:
                                bledy.append(f'Wiersz {i}: brak nazwy leku')
                                continue

                            Lek.objects.get_or_create(
                                nazwa=nazwa,
                                uzytkownik=request.user,
                                defaults={
                                    'substancja_aktywna': substancja,
                                    'instrukcja': instrukcja,
                                }
                            )
                            sukces += 1

                except Exception as e:
                    bledy.append(f'Błąd odczytu pliku: {e}')

            if not bledy:
                return redirect('strona_glowna')
    else:
        form = ImportCSVForm()

    return render(request, 'leki/import_csv.html', {
        'form': form,
        'bledy': bledy,
        'sukces': sukces,
        'tytul': 'Importuj leki z CSV',
    })


from django.core.paginator import Paginator

@login_required
def lista_lekow(request):
    # --- Filtrowanie ---
    nazwa = request.GET.get('nazwa', '')
    substancja = request.GET.get('substancja', '')
    
    leki = Lek.objects.filter(uzytkownik=request.user)
    
    if nazwa:
        leki = leki.filter(nazwa__icontains=nazwa)
    if substancja:
        leki = leki.filter(substancja_aktywna__icontains=substancja)

    # --- Stronicowanie ---
    na_strone = int(request.GET.get('na_strone', 5))
    paginator = Paginator(leki, na_strone)
    strona = request.GET.get('strona', 1)
    leki_strona = paginator.get_page(strona)

    return render(request, 'leki/lista_lekow.html', {
        'leki': leki_strona,
        'paginator': paginator,
        'nazwa': nazwa,
        'substancja': substancja,
        'na_strone': na_strone,
    })


@login_required
def lista_przyjec(request):
    # --- Filtrowanie ---
    nazwa_leku = request.GET.get('nazwa_leku', '')
    status = request.GET.get('status', '')
    data_od = request.GET.get('data_od', '')
    samopoczucie = request.GET.get('samopoczucie', '')

    przyjecia = PrzyjecieLeku.objects.filter(uzytkownik=request.user).select_related('lek')

    if nazwa_leku:
        przyjecia = przyjecia.filter(lek__nazwa__icontains=nazwa_leku)
    if status:
        przyjecia = przyjecia.filter(status=status)
    if data_od:
        przyjecia = przyjecia.filter(data_godzina__date__gte=data_od)
    if samopoczucie:
        przyjecia = przyjecia.filter(notatka_samopoczucia__icontains=samopoczucie)

    # --- Stronicowanie ---
    na_strone = int(request.GET.get('na_strone', 5))
    paginator = Paginator(przyjecia, na_strone)
    strona = request.GET.get('strona', 1)
    przyjecia_strona = paginator.get_page(strona)

    return render(request, 'leki/lista_lekow.html', {
        'leki': leki_strona,
        'page_obj': leki_strona,  # ← dodaj
        'paginator': paginator,
        'nazwa': nazwa,
        'substancja': substancja,
        'na_strone': na_strone,
    })

    return render(request, 'leki/lista_przyjec.html', {
        'przyjecia': przyjecia_strona,
        'page_obj': przyjecia_strona,  # ← dodaj
        'paginator': paginator,
        'nazwa_leku': nazwa_leku,
        'status': status,
        'data_od': data_od,
        'samopoczucie': samopoczucie,
        'na_strone': na_strone,
    })
    
    
